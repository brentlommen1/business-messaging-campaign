import telnyx
import time
import configparser
import openpyxl
import requests
import boto3
import os
import json
from urllib.parse import urlunsplit, urlparse
from botocore.exceptions import ClientError
from flask import Flask, request
from threading import Thread


num_of_messages_per_min = 6

throttle_time = 60/num_of_messages_per_min

config_parser = configparser.RawConfigParser()
config_parser.read_file(open(r'config.txt'))
your_telnyx_numbers = config_parser.get("config", "telnyx_numbers").split(",")
approved_number = config_parser.get("config", "approved_number")
excel_file_path = config_parser.get("config", "excel_file_path")
test_mode = config_parser.get("config", "test_mode").lower() == "true"
test_numbers = config_parser.get("config", "test_numbers")
TELNYX_MMS_S3_BUCKET = "cannanorth"

is_processing = False

app = Flask(__name__)

def download_file(url):
    r = requests.get(url, allow_redirects=True)
    file_name = os.path.basename(urlparse(url).path)
    open(file_name, "wb").write(r.content)
    return file_name

def upload_file(file_path):
    global TELNYX_MMS_S3_BUCKET
    s3_client = boto3.client("s3")
    file_name = os.path.basename(file_path)
    try:
        extra_args = {
            "ContentType": "application/octet-stream",
            "ACL": "public-read"
        }
        s3_client.upload_file(
            file_path,
            TELNYX_MMS_S3_BUCKET,
            file_name,
            ExtraArgs=extra_args)
    except ClientError as e:
        print("Error uploading file to S3")
        print(e)
        quit()
    return f"https://{TELNYX_MMS_S3_BUCKET}.s3.amazonaws.com/{file_name}"

def media_downloader_uploader(url):
    file_location = download_file(url)
    file_url = upload_file(file_location)
    return file_url


def readExcelFile():
    destination_numbers = []

    try:
        print("Reading excel file from " + excel_file_path)
        dataframe = openpyxl.load_workbook(excel_file_path)
        dataframe1 = dataframe.active

        stop_requested_col_num = 0
        for col in dataframe1.iter_cols(1, dataframe1.max_row):
            header_cell = col[0].value
            if header_cell == "stop_requested":
                break
            stop_requested_col_num = stop_requested_col_num + 1

        col_num = 0
        for col in dataframe1.iter_cols(1, dataframe1.max_row):
            header_cell = col[0].value
            if header_cell == "mobile_phone":
                break
            col_num = col_num + 1

        for row in dataframe1.iter_rows(1, dataframe1.max_row):
            cell_value = row[col_num].value
            if cell_value == "mobile_phone":
                continue
            if cell_value != "" and cell_value is not None :
                phone_number = str(int(cell_value))
                if phone_number[0] != "+1":
                    phone_number = "+1" + phone_number
                    stop_requested = row[stop_requested_col_num].value
                    if(stop_requested != "TRUE"):
                        destination_numbers.append(phone_number)
                    else:
                        print("removing number " + phone_number)
        print("Successfully read " + str(len(destination_numbers)) + " numbers from file")

    except FileExistsError as e:
        print(e)

    if test_mode:
        print("Test mode enabled, replacing client numbers with the test numbers: " + str(test_numbers))
        destination_numbers = [test_numbers]
    return destination_numbers

def markNumbersBlocked(blocked_nums):
    dataframe = openpyxl.load_workbook(excel_file_path)
    dataframe1 = dataframe.active

    for blocked_num in blocked_nums:

        col_num = 0
        for col in dataframe1.iter_cols(1, dataframe1.max_row):
            header_cell = col[0].value
            if header_cell == "stop_requested":
                break
            col_num = col_num + 1

        mobile_num_col = 0
        for col in dataframe1.iter_cols(1, dataframe1.max_row):
            header_cell = col[0].value
            if header_cell == "mobile_phone":
                break
            mobile_num_col = mobile_num_col + 1

        row_num = 0
        for row in dataframe1.iter_rows(1, dataframe1.max_row):
            cell_value = row[mobile_num_col].value
            if cell_value == "mobile_phone":
                row_num = row_num + 1
                continue
            if cell_value != "" and cell_value is not None:
                phone_number = str(int(cell_value))
                if phone_number[0] != "+1":
                    phone_number = "+1" + phone_number
                    if phone_number == blocked_num:
                        break
            row_num = row_num + 1
        dataframe1.cell(row=row_num+1, column=col_num+1, value="TRUE")
    dataframe.save(excel_file_path)





def sendMessageToAll(message, destination_numbers, media_url):
    count = 0
    pool_size = len(your_telnyx_numbers)
    sleep_time = throttle_time / pool_size
    message_ids = []
    blocked_numbers = []
    for dest_number in destination_numbers:
        num_position = count % pool_size
        from_number = your_telnyx_numbers[num_position]
        try:
            response = telnyx.Message.create(
                from_=from_number,
                to=dest_number,
                text=message,
                media_urls=media_url,
                use_profile_webhooks=False
            )
            message_id = response['id']
            message_ids.append(message_id)

        except Exception as e:
            error_message = e._message
            if "40300" in error_message:
                print("Message blocked due to STOP request... marking number " + str(dest_number) + " blocked.")
                blocked_numbers.append(dest_number)


        if count % 1000 == 0:
            if count == 0:
                print("Starting to send message to " + str(len(destination_numbers)) + " clients")
            else:
                print(str(count) + " messages sent.")
        count = count + 1
        time.sleep(sleep_time)
    markNumbersBlocked(blocked_numbers)
    print("DONE SENDING MESSAGES")



@app.route('/webhooks', methods=['POST'])
def webhooks():
    if not is_processing:
        body = request.json
        thread = Thread(target=processPostRequest, args=(body,))
        thread.start()
    return '', 200


def processPostRequest(body):
    global is_processing
    is_processing = True
    data = body['data']
    payload = data['payload']
    phone_number = payload['from']['phone_number']
    text_message = payload['text']
    direction = payload['direction']
    medias = payload["media"]

    media_urls = list(map(lambda media: media_downloader_uploader(media["url"]), medias))

    if phone_number == approved_number and direction == "inbound":
        destination_numbers = readExcelFile()
        sendMessageToAll(text_message, destination_numbers, media_urls)

        success_message = "Finished Sending Message"
        telnyx.Message.create(
            from_=your_telnyx_numbers[0],
            to=approved_number,
            text=success_message,
        )
    is_processing = False


def serverThread():
    app.run(port=5000)


def main():
    server_thread = Thread(target=serverThread)
    server_thread.start()


if __name__ == "__main__":
    main()
